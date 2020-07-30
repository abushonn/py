
# copy WCS-LearningProgress-master/data/wcs_test_questions_empty.xlsx, to wcs_test_questions.xlsx
# edit wcs_test_questions.xlsx:
# fill in intent examples from test set: sheet 'TestQuestions', columns 'Question Text', 'Intent', 'Reference'
# Reference should be the dataframe index of the example (pointing to the original df containing training+test sets)

import os
from shutil import copyfile
#import openpyxl
from openpyxl import load_workbook
import pandas as pd

def copy_file(folder_name, source_file_name, destination_file_name):
    '''
    Copy to destination_file_name within the same folder_name.
    :param folder_name:
    :param source_file_name:
    :param destination_file_name:
    :return:
    '''
    src = os.path.join(folder_name, source_file_name)
    dst = os.path.join(folder_name, destination_file_name)
    copyfile(src, dst)

def copy_file(source_folder_name, source_file_name, destination_folder_name, destination_file_name):
    '''
    Copy to files residing in different folders.
    :param source_folder_name:
    :param source_file_name:
    :param destination_file_name:
    :param destination_file_name:
    :return:
    '''
    src = os.path.join(source_folder_name, source_file_name)
    dst = os.path.join(destination_folder_name, destination_file_name)
    copyfile(src, dst)

def write_intent_examples(df_intent_examples, filepath):
    '''
    fill in intent examples in sheet 'TestQuestions', columns 'Question Text' ('A'), 'Intent' ('B'), 'Reference' ('C')
    :param df_intent_examples: pandas.DataFrame with columns "example", "intent" 
    :param filepath: string, path to xlsx file to edit
    :return: 
    '''

    # open xlsx file, sheet 'TestQuestions'
    wb = load_workbook(filepath)
    # Open a EXISTING worksheet in the workbook
    sheet = wb.get_sheet_by_name('TestQuestions')
    rownum = 2 # start writing from second row, first row is the header
    for index, row in df_intent_examples.iterrows():
        # write 'example' in column 'Question Text' ('A')
        cell_str = 'A' + str(rownum)
        sheet[cell_str] = row['example']
        # write 'intent' in column 'Intent' ('B')
        cell_str = 'B' + str(rownum)
        sheet[cell_str] = row['intent']
        # write index in column 'Reference' ('C')
        cell_str = 'C' + str(rownum)
        sheet[cell_str] = index

        rownum += 1

    wb.save(filepath)

def read_cell_value(filepath, sheet_name, cell_str):
    # note: each auto-generated Excel file needs to be opened and saved manually (or with some utility) in order for the formulas to be calculated and for the values to be saved.
    # otherwise reading the value of formula cells will return None.
    wb = load_workbook(filepath, data_only = True) # to read values rather than formulas
    sheet = wb.get_sheet_by_name(sheet_name)
    return sheet[cell_str].value

def read_sheet_to_df(filepath, sheet_name):
    return pd.read_excel(filepath, sheet_name = sheet_name)

def write_df_to_sheet(df, filepath, sheet_name):
    df.to_excel(filepath, sheet_name = sheet_name, index = False)

if __name__ == "__main__":
    folder_name = 'WCS-LearningProgress-master\\data'
    source_file_name = 'wcs_test_questions_empty.xlsx'
    destination_file_name = 'wcs_test_questions.xlsx'

    copy_file(folder_name, source_file_name, destination_file_name)