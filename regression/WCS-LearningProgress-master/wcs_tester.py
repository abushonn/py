# -*- coding: utf-8 -*-
#!/usr/bin/python
'''
This code was developed for reporting purposes to check if Watson Conversation
Service / Watson Assistant correctly maps questions inputted by the user to
the expected intents.  Use this code to track continuous improvement of your
chatbot system.

N.B. Before you can run this code, make sure you have the data/Results.xlsx
file in the same directory as where the script WCS_Tester.py exists.

In order to use this code, please select a certain number of questions (say 250)
at random from the collected Questions and paste them in the TestQuestions
worksheet of the data/wcs_test_questions.xlsx file.

Then run this code in your local Command Line Tool (Terminal for Macs or Command
Prompt for Windows). Once the code runs its course, you will be notified on the
CLI with a success message and your Results.xlsx file will be updated with the
latest results.

Usage: wcs_tester.py username password workspace_id

Be sure to create a copy of the updated Results.xlsx file and
rename the copy to a standard name of your choice such as WCS_Results_v1.xlsx.
You will require the Results.xlsx file for future runs.

Before you can run this code, please be sure to read the ReadMe-LearningScript.docx
file as it contains information on configuration and installation before this code can
run. You will require:
1. python 3
2. Python installation Packager (pip)
3. Openpyxl
4. xlrd
5. Watson developer cloud

Code Developed by: Shoaib Bilal (IBM Australia)
Initial Version Date: 25/06/2016
Current Version: 1.4
Last updated on: 21/05/2018
Code Modified by: Sue Ann Chen (IBM Research Australia), Jason Jingshi Li (IBM Australia), Jaysen Ollerenshaw
'''
import xlrd, warnings
import csv, os, re
import datetime, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from watson_developer_cloud import AssistantV1

'''------------------------------------------------------------------'''
''' Following are a set of global variables. '''

api_version = "2018-07-10"

WA_URL = "https://gateway.watsonplatform.net/assistant/api"

data_dir = "./data"
INPUT_FILE = "wcs_test_questions.xlsx"
input_sheet = "TestQuestions"
temp_file = "WCS_Mapping.csv"
OUTPUT_FILE = "wcs_test_results.xlsx"
output_sheet = "WCS Results"

# The input spreadsheet containing test questions.
book = None
# The question worksheet to load data from.
sh = None

title_qid = "Reference"
title_question = "Question Text"
title_intent = "Intent"

title_entityname = "Entity Name"
title_entityvalue = "Entity Value"

title_status = "Status"
title_inscope = "Scope"

title_answer = "Expected Answer"
answer_bubble_separator = " "

col_question = None
col_qid = None
col_intent = None
col_entityname = None
col_entityvalue = None
col_inscope = None
col_status = None
col_answer = None

api_context = {}

scope_keys = ['Critical',"[On]\-*[topic]",'System',"[Chit]\-*[Chat]"]
scope_prefixes = ['X.','O.','S.','C.']
problem_chars = ['\'','\,',' ']
irrelevant_threshold = 0.2
unsure_threshold = 0.5
confident_threshold = 0.8
very_confident_threshold = 0.95

'''------------------------------------------------------------------'''
''' Utility functions. '''

def strip_non_ascii(string):
    ''' This cleans the questions variable of any non-ascii characters such as
    smart apostrophes and en-dashes '''

    # Strip question of non-ascii characters
    stripped = (c for c in string if 0<ord(c)<127)
    # return question stripped of non-ascii characters.
    return ''.join(stripped)

def ReadQuestion(rownum):
    ''' This function reads the question text from a cell in the given row
    in the TestQuestions worksheet.'''

    question = sh.cell_value(rowx=rownum, colx=col_question)
    # Merge any whitespace and strip non-ascii before returning.
    return strip_non_ascii(re.sub('\s+',' ',question))

def init_question_sheet(sheet):
    ''' Open input Excel file and initialise column numbers based
    on headings in source spreadsheet. '''
    global book
    global sh
    global col_qid
    global col_question
    global col_intent
    global col_entityname
    global col_entityvalue
    global col_status
    global col_inscope
    global col_answer
    
    # Open the excel file QAMaster and assign it the variable book
    book = xlrd.open_workbook(data_dir + "/" + INPUT_FILE)
    # Open the sheet "TestQuestions" and assign it the variable sh
    sh = book.sheet_by_name(input_sheet)

    col_qid = value_from_key(sh, title_qid)
    col_question = value_from_key(sh, title_question)
    col_intent = value_from_key(sh, title_intent)
    col_entityname = value_from_key(sh, title_entityname)
    col_entityvalue = value_from_key(sh, title_entityvalue)
    col_status = value_from_key(sh, title_status)
    col_inscope = value_from_key(sh, title_inscope)
    col_answer = value_from_key(sh, title_answer)

    print ('Discovered columns: ',
        title_qid, '=', col_qid, ', ',
        title_question, '=', col_question, ', ',
        title_intent, '=', col_intent, ', ',
        title_entityname, '=', col_entityname, ', ',
        title_entityvalue, '=', col_entityvalue, ', ',
        title_status, '=', col_status, ', ',
        title_inscope, '=', col_inscope, ', ',
        title_answer, '=', col_answer, '.')
    if col_question is None or col_intent is None or col_entityname is None or col_entityvalue is None:
        print ('Mandatory columns not found. Please check column names in spreadsheet.')
        sys.exit

def createnewdict(responsedict):
    newdict = {"text": responsedict.get('input').get('text'), \
                "classifier_id": responsedict.get('context').get('conversation_id'), \
                "classes": [{"class_name": responsedict.get("intents")[idx].get('intent'), \
                            "confidence": responsedict.get("intents")[idx].get('confidence')} \
                            for idx in range(len(responsedict.get("intents")))],
                "entities": [{"entity_name": responsedict.get("entities")[index].get('entity'),\
                            "entity_value": responsedict.get("entities")[index].get('value')}
                            for index in range(len(responsedict.get("entities")))],
                "output": answer_bubble_separator.join(responsedict.get('output').get('text'))
                    }
    return newdict


def get_assistant():
    username = sys.argv[1]
    password = sys.argv[2]

    assistant = AssistantV1(
        version=api_version,
        username=username,
        password=password,
        url=WA_URL
    )

    return assistant

def preapi_call():
    ''' Call Conversation with an empty request to get the test started.'''
    global api_context

    wa_asistant = get_assistant()

    context = {}
    workspace_id = sys.argv[3]
    response = wa_asistant.message(workspace_id = workspace_id, input = {'text':''}, context = context)
    api_context = response.get_result().get('context')
    return

def api_call(question):
    ''' This function makes the api call by sending the question to WCS and
    getting a response which is stored in the variable classes.'''

    global api_context

    wa_asistant = get_assistant()
    workspace_id = sys.argv[3]

    response = wa_asistant.message(workspace_id = workspace_id,
                                        input = {'text':question},
                                        alternate_intents = True,
                                        context = api_context,
                                        entities = None,
                                        intents = None,
                                        output = None,
                                        nodes_visited_details = None,
                                        headers = {'X-Watson-Learning-Opt-Out': 'true'})
    api_context = response.get_result().get('context')
    classes = createnewdict(response.get_result())
    print ('CLASSES:')
    print (classes)
    return classes

def ReadQID(rownum):
    ''' This function reads the question id for the given row from the
    TestQuestions worksheet. '''

    if col_qid != None:
        Qid = sh.cell_value(rowx=rownum, colx=col_qid)
        return Qid
    return ''

def ReadIntent(rownum):
    ''' This function reads the Intent for the given row from the
    TestQuestions worksheet.  Invalid chars are replaced by underscores.
    Prefixes are added based on the scope column. '''

    def __replace_punc(intent,changelist):
        # Find and replace spaces and punctuation with underscores
        NewIntentString = intent.rstrip('?')
        for item in changelist:
            NewIntentString = NewIntentString.replace(item, '_')
        return NewIntentString

    def __add_prefix(intent, scope_val, scope_handle):
        for handle_now, prefix_now in scope_handle:
            if re.search(handle_now, scope_val, re.IGNORECASE):
                intent = prefix_now + intent
        return intent

    # cell value of Question Intent column stored in variable Intent.
    Intent = sh.cell_value(rowx=rownum, colx=col_intent).rstrip()
    Intentv2 = __replace_punc(Intent, problem_chars)
    if col_inscope != None:
        scope_val = InScope(rownum)
        scope_handle = zip(scope_keys, scope_prefixes)
        Intentv3 = __add_prefix(Intentv2, scope_val, scope_handle)
        return Intentv3
    return Intentv2

def GetStatus(rownum):
    ''' This function reads the status for a given row from the
    TestQuestions worksheet. '''
    if col_status != None:
        status = sh.cell_value(rowx=rownum, colx=col_status)
        return status
    return ''

def InScope(rownum):
    ''' This function reads the Scope for a given row from the
    TestQuestions worksheet. '''
    if col_inscope != None:
        scope = sh.cell_value(rowx=rownum, colx=col_inscope)
        return scope
    return ''

def GetExpectedAnswer(rownum):
    ''' This function reads the Expected Answer for a given row from the
    TestQuestions worksheet.
    Answers are optional, and used for regression testing. '''
    if col_answer != None:
        ans = sh.cell_value(rowx=rownum, colx=col_answer)
        return ans
    return ''

def value_from_key(sheet, key):
    ''' Works out which column in the spreadsheet has the given heading. '''
    titles = sheet.row(0)  # 1st row
    headers = [str(cell.value) for cell in titles]
    for col_index in range(sheet.ncols):
        if headers[col_index] == key:
            return col_index

def Convert_To_Excel():
    ''' This following code converts a csv created in main(), to an ALREADY
    EXISTING xlsx file format so we can perform standardised testing on it'''

    # Open the csv file called NLC_Mapping.csv
    OpenFile  = open(r'./data/WCS_Mapping.csv')
    # Define the dialect variable of the csv file as 'comma'
    csv.register_dialect('comma', delimiter=',')
    # Read the csv file delimited by commas
    reader = csv.reader(OpenFile, dialect='comma')

    # Ignore warning messages
    warnings.simplefilter("ignore")
    # Load a PREVIOUSLY EXISTING excel workbook of results
    wb = load_workbook("./data/wcs_test_results.xlsx")
    # Open a EXISTING worksheet in the workbook by the name 'WCS Mapping'
    sheet = wb.get_sheet_by_name('WCS Mapping')

    # For each cell in each row the csv file (reader)
    for row_index, row in enumerate(reader):
        # for each cell in each column in the csv file (reader)
        for column_index, cell in enumerate(row):
            # Get each cell in column in csv file and assign it the variable
            # column_letter
            column_letter = get_column_letter((column_index + 1))
            # Populate column_letter in the excel file
            sheet.cell('%s%s'%(column_letter, (row_index + 1))).value = convertCell(cell)
    # Close the CSV file to prevent any unnecassary memory leakage
    OpenFile.close()
    # Return the excel file as a result.
    return wb

def convertCell(value):
    ''' Convert the string value of a CSV entry to the correct format for Excel. '''
    try:
        float_value = float(value)
        int_value = int(float_value)
        if int_value == float_value:
            return int_value
        else:
            return float_value
    except ValueError:
        return value

'''------------------------------------------------------------------'''

def matchEntities(entityName, entityValue, entitiesList):
    if entityName == "" or entityName == None:
        return True
    for entity in entitiesList:
        if "entity_name" not in entity:
            continue
        if entityName.lower() == entity["entity_name"].lower():
            if entityValue == "" or entityValue == None:
                return True
            if "entity_value" not in entity:
                continue
            if entityValue.lower() == entity["entity_value"].lower():
                return True
    return False

def getEntitiesString(entitiesList):
    returnString = ""
    for entity in entitiesList:
        if "entity_name" not in entity:
            continue
        entityName = entity["entity_name"]
        if returnString == "":
            returnString = "@" + entityName
        else:
            returnString = returnString + "|@" + entityName
        if "entity_value" not in entity:
            continue
        entityValue = entity["entity_value"]
        if " " in entityValue:
            returnString = returnString+ ":(" + entityValue + ")"
        else:
            returnString = returnString + ":" + entityValue
    return returnString

'''------------------------------------------------------------------'''

def main():
    
    # Confirm minimum command line arguments.
    if len(sys.argv) <= 4:
        print ('usage: wcs_tester.py <username> <password> <workspace_id> <wa_url>')
        sys.exit

    WA_URL = sys.argv[4]
    # record start time
    startTime = datetime.datetime.now()

    init_question_sheet(sh)

    # Create a csv file by the name WCS_Mapping.csv
    outputFile = open(data_dir + "/" + temp_file, "w+", newline='')
    # Call csv writer module and assign it the variable f
    f = csv.writer(outputFile)

    # Now write the column header line in the csv file.
    # Take all the optional fields into account.
    headers = [title_question]
    if col_qid != None:
        headers.append(title_qid)
    if col_status != None:
        headers.append(title_status)
    if col_inscope != None:
        headers.append(title_inscope)
    headers.extend(["Expected Intent", "Entity Name", "Entity Value",])
    if col_answer != None:
        headers.extend(["Answer", "Expected Answer Matched"])
    headers.extend(["Intent 1", "Confidence 1",
                "Intent 2", "Confidence 2",
                "Intent 3", "Confidence 3",
                "Intent 4", "Confidence 4",
                "Intent 5", "Confidence 5",
                "Entities Found", "Expected Entity Matched"])
    headers.extend(["Top Result", "Top 5", "Confusion",
                "Recommendation","Recommended Intent"])
    f.writerow(headers)

    preapi_call()

    # Loop through all the input question rows in the source spreadsheet.
    for row_number in range(1, sh.nrows):

        print ("Processing row " + str(row_number))

        # Get the test Question for this row.
        a = ReadQuestion(row_number)
        print (a)
        # Call Watson to see what intent / entities are matched.
        b = api_call(a)

        # What was the expected intent from the input sheet?
        intentString = ReadIntent(row_number)

        # Read in entity name and values
        entityName = sh.cell_value(rowx=row_number, colx=col_entityname).replace(" ", "_")
        entityValue = sh.cell_value(rowx=row_number, colx=col_entityvalue)

        print ("Looking for entity name: " + entityName + ", entity value: " + entityValue)
        if not matchEntities(entityName, entityValue, b["entities"]):
            print (entityName + " " + entityValue + " not found in" + str(b["entities"]))
        else:
            print (entityName + " " + entityValue + " found.")

        # Determine some recommendations based on the WCS results.

        top_result = ""
        if (intentString > ""):
            if (b["classes"][0]["confidence"] >= irrelevant_threshold):
                if (intentString == b["classes"][0]["class_name"]):
                    top_result = "True Positive"
                else:
                    top_result = "False Positive"
            else:
                top_result = "False Negative"
        else:
            if (b["classes"][0]["confidence"] < irrelevant_threshold):
                top_result = "True Negative"
            else:
                top_result = "False Positive"

        top_five = "FALSE"
        if (((b["classes"][0]["class_name"] == intentString) and
                (b["classes"][0]["confidence"] >= irrelevant_threshold)) or
                   ((b["classes"][1]["class_name"] == intentString) and
                (b["classes"][1]["confidence"] >= irrelevant_threshold)) or
                   ((b["classes"][2]["class_name"] == intentString) and
                (b["classes"][2]["confidence"] >= irrelevant_threshold)) or
                   ((b["classes"][3]["class_name"] == intentString) and
                (b["classes"][3]["confidence"] >= irrelevant_threshold)) or
                   ((b["classes"][4]["class_name"] == intentString) and
                (b["classes"][4]["confidence"] >= irrelevant_threshold))) :
            top_five = "TRUE"

        # Crude calculation of confusion.  Don't look!
        confusion = ""
        if (b["classes"][0]["confidence"] > b["classes"][1]["confidence"] * 1.5 + 0.1):
            confusion = "Low"
        elif (b["classes"][0]["confidence"] > b["classes"][1]["confidence"] * 1.3 + 0.05):
            confusion = "Med"
        elif (b["classes"][0]["confidence"] > b["classes"][2]["confidence"] * 1.3 + 0.05):
            confusion = "High"
        else:
            confusion = "Extreme"
            
        # Did we get the right answer?
        returned_answer = None
        answer_correct = None
        if col_answer != None:
            returned_answer = b["output"]
            answer_correct = returned_answer == GetExpectedAnswer(row_number)

        # Determine a recommendation for this example input.
        if (b["classes"][0]["confidence"] == 1.0) and (b["classes"][1]["confidence"] == 1.0):
            # This example is already listed in two different intents.
            recommendation = "Remove Duplicate"
        elif intentString > "" :
            # We have an intent that we want it to map to.  How well did that go?
            if (b["classes"][0]["class_name"] == intentString) :
                if (b["classes"][0]["confidence"] < irrelevant_threshold) :
                    # This example does not match to any intent.
                    recommendation = "Irrelevant"
                elif (b["classes"][0]["confidence"] < unsure_threshold) :
                    # This example matches the expected intent, but was unconfident.
                    recommendation = "Include"
                elif (b["classes"][0]["confidence"] < confident_threshold) :
                    # This example matches the expected intent, but was not highly confident.
                    recommendation = "Include"
                elif (b["classes"][1]["confidence"] >= confident_threshold) :
                    # Top two intents are confident.
                    recommendation = "Switch"
                elif (b["classes"][0]["confidence"] < very_confident_threshold) :
                    # This example matches the expected intent confidently.
                    recommendation = "Accurate"
                else :
                    # This example is matched very confidently to the expected intent.
                    recommendation = "Very Accurate"
            else :
                if (b["classes"][0]["confidence"] < irrelevant_threshold) :
                    # This example does not match to any intent.
                    recommendation = "Irrelevant"
                elif (b["classes"][0]["confidence"] < unsure_threshold) :
                    # This example didn't match the expected intent, and the matched intent was unconfident.
                    recommendation = "Include"
                elif (b["classes"][0]["confidence"] < confident_threshold) :
                    # This example didn't match the expected intent, and the matched intent was medium confidence.
                    recommendation = "Include"
                elif (b["classes"][0]["confidence"] < very_confident_threshold) :
                    # This example didn't match the expected intent, and the matched intent was confident.
                    recommendation = "Train"
                else :
                    # This example didn't match the expected intent, maybe the other intent is a better match.
                    recommendation = "Train"
        # We don't know what it should map to.  What should it be?
        elif b["classes"][1]["confidence"] >= very_confident_threshold :
            # Top two intents are both very confident.  This indicates possible confusion.
            recommendation = "Confused"
        elif b["classes"][1]["confidence"] >= confident_threshold :
            # Top two intents are both confident.
            recommendation = "Switch"
        elif b["classes"][0]["confidence"] < irrelevant_threshold :
            # No intent seems to match this example.
            recommendation = "Irrelevant"
        elif b["classes"][0]["confidence"] < unsure_threshold :
            # This example is matched, but with low confidence.
            recommendation = "Train"
        elif b["classes"][0]["confidence"] < confident_threshold :
            # This example is matched, but not hugely confident.
            recommendation = "Train"
        elif b["classes"][0]["confidence"] < very_confident_threshold :
            # This example is matched fairly confidently.
            recommendation = "Accurate"
        else :
            # This example is matched very confidently to an intent.
            recommendation = "Very Accurate"

        recommended_intent = "";
        if (recommendation == "Train") :
            recommended_intent = b["classes"][0]["class_name"]
        elif (recommendation == "Switch") :
            recommended_intent = b["classes"][1]["class_name"]
        elif (recommendation == "Include") :
            recommended_intent = intentString
        elif (recommendation == "Accurate" and intentString > "") :
            recommended_intent = b["classes"][0]["class_name"]

        # Populate the row with original data plus results from WCS along with the recomendations.

        output_row = ([b["text"]])
        if col_qid != None:
            output_row.append(ReadQID(row_number))
        if col_status != None:
            output_row.append(GetStatus(row_number))
        if col_inscope != None:
            output_row.append(InScope(row_number))
        output_row.extend([intentString,
                    entityName,
                    entityValue,])
        if col_answer != None:
            output_row.extend([returned_answer, "TRUE" if answer_correct else "FALSE"])
        output_row.extend([b["classes"][0]["class_name"],
                    b["classes"][0]["confidence"],
                    b["classes"][1]["class_name"],
                    b["classes"][1]["confidence"],
                    b["classes"][2]["class_name"],
                    b["classes"][2]["confidence"],
                    b["classes"][3]["class_name"],
                    b["classes"][3]["confidence"],
                    b["classes"][4]["class_name"],
                    b["classes"][4]["confidence"],
                    getEntitiesString(b["entities"]),
                    "TRUE" if matchEntities(entityName, entityValue, b["entities"]) else "FALSE"])
        output_row.extend([top_result,
                    top_five,
                    confusion,
                    recommendation,
                    recommended_intent,])

        f.writerow (output_row)
        currentTime = datetime.datetime.now()
        timeElapsed = currentTime - startTime
        print ("Time elapsed: " + str(timeElapsed) + ", average " + str(timeElapsed/row_number) + " per question.")
    # Close the created CSV file.
    outputFile.close()

    # Call the function to convert to csv to an excel file
    d = Convert_To_Excel()

    # Open the already existing sheet WCS Mapping in the excel file.
    ws = d.get_sheet_by_name('WCS Mapping')

    # Apply filters on the column headers.
    ws.auto_filter.ref = "A:W"
    # Delete the CSV file. We no longer need it.
    #os.remove("./data/WCS_Mapping.csv")
    # Save the amended excel file.
    d.save('./data/wcs_test_results.xlsx')
    
    print ("Done! You can now check your updated file for the latest results.")
    return

if __name__ == '__main__':
    main()
